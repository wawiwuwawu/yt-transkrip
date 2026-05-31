import yt_dlp
import whisper
import pandas as pd
import os
import torch
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def format_waktu(detik):
    """Mengubah detik menjadi format HH:MM:SS atau MM:SS"""
    hours = int(detik // 3600)
    minutes = int((detik % 3600) // 60)
    seconds = int(detik % 60)
    
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    else:
        return f"{minutes:02d}:{seconds:02d}"

def create_unique_output_folder(video_title="video", base_folder="transcripts"):
    """
    Membuat folder output unik berdasarkan timestamp dan judul video
    """
    # Get current timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Ultra-safe title cleaning - hanya alphanumeric dan spasi
    safe_title = re.sub(r'[^a-zA-Z0-9\s]', '', video_title)  # Only alphanumeric and spaces
    safe_title = re.sub(r'\s+', '_', safe_title)  # Replace spaces with underscores
    safe_title = safe_title.strip('_')[:25]  # Limit to 25 chars
    if not safe_title:
        safe_title = "video"
    
    # Create folder name: transcripts/20231222_143045_VideoTitle
    folder_name = f"{timestamp}_{safe_title}"
    
    # Ensure base folder exists first
    try:
        if not os.path.exists(base_folder):
            os.makedirs(base_folder, exist_ok=True)
            print(f"📁 Created base folder: {base_folder}")
        
        output_folder = os.path.join(base_folder, folder_name)
        os.makedirs(output_folder, exist_ok=True)
        print(f"📁 Output folder created: {output_folder}")
        
        # Verify folder was created successfully
        if not os.path.exists(output_folder):
            raise Exception(f"Failed to create folder: {output_folder}")
            
        return output_folder
        
    except Exception as e:
        print(f"❌ Error creating folder: {e}")
        # Ultra fallback - just use timestamp
        fallback_folder = os.path.join(base_folder, timestamp)
        try:
            os.makedirs(fallback_folder, exist_ok=True)
            print(f"📁 Using fallback folder: {fallback_folder}")
            return fallback_folder
        except:
            # Last resort - current directory with timestamp
            final_fallback = f"output_{timestamp}"
            os.makedirs(final_fallback, exist_ok=True)
            print(f"📁 Using final fallback: {final_fallback}")
            return final_fallback

def download_audio_ytdlp(video_url, output_path="temp_audio"):
    """Download audio menggunakan yt-dlp"""
    ydl_opts = {
        'format': 'bestaudio/best',
        'extractaudio': True,
        'audioformat': 'mp3',
        'outtmpl': f'{output_path}.%(ext)s',
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }],
        'quiet': True,  # Mengurangi output verbose
    }
    
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            # Get video info
            info = ydl.extract_info(video_url, download=False)
            video_info = {
                'title': info.get('title', 'Unknown'),
                'duration': info.get('duration', 0),
                'uploader': info.get('uploader', 'Unknown'),
                'view_count': info.get('view_count', 0),
                'upload_date': info.get('upload_date', 'Unknown')
            }
            
            # Download audio
            ydl.download([video_url])
            
            # Find the downloaded file
            audio_file = None
            for ext in ['mp3', 'wav', 'm4a', 'webm']:
                potential_file = f"{output_path}.{ext}"
                if os.path.exists(potential_file):
                    audio_file = potential_file
                    break
            
            if not audio_file:
                raise Exception("Audio file tidak ditemukan setelah download")
                
            return audio_file, video_info
            
    except Exception as e:
        print(f"Error saat download: {e}")
        return None, None

def create_beautiful_table_excel(video_info, segments, paragraphs, filename, output_folder="."):
    """Membuat Excel dengan format tabel yang sangat rapi dan professional"""
    try:
        # Ensure output folder exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder, exist_ok=True)
            print(f"📁 Created missing output folder: {output_folder}")
        
        # Ensure full path
        full_filename = os.path.join(output_folder, filename)
        print(f"📊 Creating Excel file: {full_filename}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transkrip Video"
        
        # Color scheme - Blue professional theme
        primary_blue = "1F4E79"
        light_blue = "D9E2F3"
        accent_blue = "2E75B6"
        white = "FFFFFF"
        light_gray = "F8F9FA"
        
        # Font definitions
        title_font = Font(name='Calibri', size=16, bold=True, color=primary_blue)
        header_font = Font(name='Calibri', size=11, bold=True, color=white)
        data_font = Font(name='Calibri', size=10)
        info_font = Font(name='Calibri', size=10, italic=True)
        
        # Fill definitions
        header_fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
        light_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
        
        # Border definitions
        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === TITLE SECTION ===
        title_cell = ws.cell(row=1, column=1, value="TRANSKRIP VIDEO YOUTUBE")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        
        # Video title
        video_title_cell = ws.cell(row=2, column=1, value=video_info['title'])
        video_title_cell.font = Font(name='Calibri', size=12, bold=True)
        video_title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.merge_cells('A2:F2')
        
        # === INFO SECTION ===
        info_row = 4
        ws.cell(row=info_row, column=1, value="Channel:").font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=info_row, column=2, value=video_info['uploader']).font = info_font
        
        ws.cell(row=info_row, column=3, value="Durasi:").font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=info_row, column=4, value=format_waktu(video_info['duration'])).font = info_font
        
        ws.cell(row=info_row, column=5, value="Tanggal:").font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=info_row, column=6, value=pd.Timestamp.now().strftime('%Y-%m-%d')).font = info_font
        
        # === TABLE HEADERS ===
        header_row = 6
        headers = ["No", "Waktu Mulai", "Waktu Selesai", "Durasi", "Kata", "Teks Transkrip"]
        column_widths = [6, 12, 12, 10, 8, 70]
        
        for col, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            ws.column_dimensions[chr(64 + col)].width = width
        
        # === TABLE DATA ===
        data_start_row = header_row + 1
        
        for i, segment in enumerate(segments):
            row = data_start_row + i
            fill = light_fill if i % 2 == 0 else PatternFill(start_color=white, end_color=white, fill_type="solid")
            
            # Calculate values
            duration = segment['end'] - segment['start']
            text = segment['text'].strip()
            word_count = len(text.split())
            
            # No
            cell = ws.cell(row=row, column=1, value=i + 1)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = fill
            
            # Start time
            cell = ws.cell(row=row, column=2, value=format_waktu(segment['start']))
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = fill
            
            # End time
            cell = ws.cell(row=row, column=3, value=format_waktu(segment['end']))
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = fill
            
            # Duration
            cell = ws.cell(row=row, column=4, value=f"{duration:.1f}s")
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = fill
            
            # Word count
            cell = ws.cell(row=row, column=5, value=word_count)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.fill = fill
            
            # Text
            cell = ws.cell(row=row, column=6, value=text)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            cell.fill = fill
            
            # Set row height based on text length
            estimated_lines = max(1, len(text) // 80)
            ws.row_dimensions[row].height = max(20, estimated_lines * 15)
        
        # === STATISTICS SHEET ===
        stats_ws = wb.create_sheet("Statistik")
        
        # Calculate statistics
        total_duration = video_info['duration']
        total_segments = len(segments)
        total_words = sum(len(seg['text'].split()) for seg in segments)
        total_chars = sum(len(seg['text']) for seg in segments)
        words_per_minute = (total_words / (total_duration / 60)) if total_duration > 0 else 0
        avg_segment_duration = total_duration / total_segments if total_segments > 0 else 0
        
        # Statistics data
        stats_data = [
            ["STATISTIK TRANSKRIP", ""],
            ["", ""],
            ["Informasi Video", ""],
            ["Judul", video_info['title']],
            ["Channel", video_info['uploader']],
            ["Durasi Total", format_waktu(total_duration)],
            ["", ""],
            ["Statistik Transkrip", ""],
            ["Total Segmen", f"{total_segments:,}"],
            ["Total Kata", f"{total_words:,}"],
            ["Total Karakter", f"{total_chars:,}"],
            ["Kata per Menit", f"{words_per_minute:.1f}"],
            ["Rata-rata Durasi Segmen", f"{avg_segment_duration:.2f} detik"],
            ["Rata-rata Kata per Segmen", f"{total_words/total_segments:.1f}"],
        ]
        
        if segments:
            durations = [seg['end'] - seg['start'] for seg in segments]
            stats_data.extend([
                ["Segmen Terpendek", f"{min(durations):.2f} detik"],
                ["Segmen Terpanjang", f"{max(durations):.2f} detik"],
            ])
        
        # Write statistics
        for row, (label, value) in enumerate(stats_data, 1):
            if row == 1:  # Title
                cell = stats_ws.cell(row=row, column=1, value=label)
                cell.font = Font(name='Calibri', size=16, bold=True, color=primary_blue)
                cell.alignment = Alignment(horizontal="center")
                stats_ws.merge_cells(f'A{row}:B{row}')
            elif label and not label.startswith("Informasi") and not label.startswith("Statistik"):  # Data
                # Label
                label_cell = stats_ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(name='Calibri', size=11, bold=True)
                label_cell.alignment = Alignment(horizontal="right")                # Value
                value_cell = stats_ws.cell(row=row, column=2, value=value)
                value_cell.font = Font(name='Calibri', size=11)
                value_cell.alignment = Alignment(horizontal="left")
            elif label:  # Section headers
                cell = stats_ws.cell(row=row, column=1, value=label)
                cell.font = Font(name='Calibri', size=12, bold=True, color=accent_blue)
                stats_ws.merge_cells(f'A{row}:B{row}')
        
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 30
        
        # === PARAGRAF SHEET ===
        if paragraphs:
            paragraf_ws = wb.create_sheet("Paragraf")
            
            # Title
            title_cell = paragraf_ws.cell(row=1, column=1, value="TRANSKRIP FORMAT PARAGRAF")
            title_cell.font = Font(name='Calibri', size=16, bold=True, color=primary_blue)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            paragraf_ws.merge_cells('A1:F1')
            
            # Video info
            video_title_cell = paragraf_ws.cell(row=2, column=1, value=video_info['title'])
            video_title_cell.font = Font(name='Calibri', size=12, bold=True)
            video_title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            paragraf_ws.merge_cells('A2:F2')
            
            # Headers untuk paragraf
            header_row = 4
            headers = ["No", "Waktu Mulai", "Waktu Selesai", "Durasi", "Kata", "Teks Paragraf"]
            column_widths = [6, 12, 12, 10, 8, 80]
            
            for col, (header, width) in enumerate(zip(headers, column_widths), 1):
                cell = paragraf_ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_border
                paragraf_ws.column_dimensions[chr(64 + col)].width = width
            
            # Data paragraf
            data_start_row = header_row + 1
            
            for i, paragraph in enumerate(paragraphs):
                row = data_start_row + i
                fill = light_fill if i % 2 == 0 else PatternFill(start_color=white, end_color=white, fill_type="solid")
                
                # No
                cell = paragraf_ws.cell(row=row, column=1, value=i + 1)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = fill
                
                # Start time
                cell = paragraf_ws.cell(row=row, column=2, value=format_waktu(paragraph['start']))
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = fill
                
                # End time
                cell = paragraf_ws.cell(row=row, column=3, value=format_waktu(paragraph['end']))
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = fill
                
                # Duration
                cell = paragraf_ws.cell(row=row, column=4, value=f"{paragraph['duration']:.1f}s")
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = fill
                
                # Word count
                cell = paragraf_ws.cell(row=row, column=5, value=paragraph['word_count'])
                cell.font = data_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                cell.fill = fill
                
                # Paragraph text
                cell = paragraf_ws.cell(row=row, column=6, value=paragraph['text'])
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
                cell.fill = fill
                
                # Set row height based on text length (lebih tinggi untuk paragraf)
                estimated_lines = max(2, len(paragraph['text']) // 100)
                paragraf_ws.row_dimensions[row].height = max(30, estimated_lines * 18)
            
            # Update statistics untuk include paragraf info
            if paragraphs:
                avg_paragraph_duration = sum(p['duration'] for p in paragraphs) / len(paragraphs)
                avg_words_per_paragraph = sum(p['word_count'] for p in paragraphs) / len(paragraphs)
                
                # Add paragraf statistics to stats sheet
                stats_data.extend([
                    ["", ""],
                    ["Statistik Paragraf", ""],
                    ["Total Paragraf", f"{len(paragraphs):,}"],
                    ["Rata-rata Durasi Paragraf", f"{avg_paragraph_duration:.1f} detik"],
                    ["Rata-rata Kata per Paragraf", f"{avg_words_per_paragraph:.1f}"],
                ])
                
                # Re-write updated statistics
                for row, (label, value) in enumerate(stats_data, 1):
                    if row <= len(stats_data):
                        if row > 14:  # Only update new rows
                            if label and not label.startswith("Statistik Paragraf") and label:  # Data
                                # Label
                                label_cell = stats_ws.cell(row=row, column=1, value=label)
                                label_cell.font = Font(name='Calibri', size=11, bold=True)
                                label_cell.alignment = Alignment(horizontal="right")
                                
                                # Value
                                value_cell = stats_ws.cell(row=row, column=2, value=value)
                                value_cell.font = Font(name='Calibri', size=11)
                                value_cell.alignment = Alignment(horizontal="left")
                            elif label == "Statistik Paragraf":  # Section header
                                cell = stats_ws.cell(row=row, column=1, value=label)
                                cell.font = Font(name='Calibri', size=12, bold=True, color=accent_blue)
                                stats_ws.merge_cells(f'A{row}:B{row}')
        
        # Save workbook with error handling
        try:
            wb.save(full_filename)
            print(f"✅ Excel file saved successfully: {full_filename}")
        except Exception as save_error:
            print(f"❌ Error saving Excel file: {save_error}")
            # Try alternative save with simplified path
            try:
                simple_filename = os.path.join(output_folder, "transkrip_backup.xlsx")
                wb.save(simple_filename)
                print(f"✅ Excel saved as backup: {simple_filename}")
            except Exception as backup_error:
                print(f"❌ Failed to save Excel backup: {backup_error}")
                # Create TXT fallback
                fallback_txt = os.path.join(output_folder, "excel_fallback.txt")
                with open(fallback_txt, 'w', encoding='utf-8') as f:
                    f.write("EXCEL CREATION FAILED - FALLBACK TEXT\n")
                    f.write("=" * 50 + "\n")
                    f.write(f"Video: {video_info['title']}\n")
                    f.write(f"Duration: {video_info['duration']}\n")
                    f.write(f"Segments: {len(segments)}\n")
                    for i, seg in enumerate(segments, 1):
                        f.write(f"{i:3d}. [{format_waktu(seg['start'])}] {seg['text'].strip()}\n")
                print(f"📝 Created text fallback: {fallback_txt}")
                
    except Exception as excel_error:
        print(f"❌ Excel creation failed: {excel_error}")
        # Create fallback text file
        fallback_file = os.path.join(output_folder, "excel_error_fallback.txt")
        try:
            with open(fallback_file, 'w', encoding='utf-8') as f:
                f.write("EXCEL CREATION ERROR - FALLBACK\n")
                f.write("=" * 40 + "\n")
                f.write(f"Error: {excel_error}\n")
                f.write(f"Video: {video_info.get('title', 'Unknown')}\n")
                f.write("Transcript data still available in other formats.\n")
            print(f"📝 Error fallback created: {fallback_file}")
        except Exception as fallback_error:
            print(f"❌ Even fallback failed: {fallback_error}")
            raise Exception(f"Critical error in Excel creation: {excel_error}")
        print(f"✅ Excel file saved successfully: {full_filename}")
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        # Create a simple text file as fallback
        fallback_file = os.path.join(output_folder, filename.replace('.xlsx', '_fallback.txt'))
        with open(fallback_file, 'w', encoding='utf-8') as f:
            f.write("TRANSKRIP VIDEO (FALLBACK)\n")
            f.write("=" * 50 + "\n")
            f.write(f"Judul: {video_info.get('title', 'Unknown')}\n")
            f.write(f"Channel: {video_info.get('uploader', 'Unknown')}\n")
            f.write("=" * 50 + "\n\n")
            for i, seg in enumerate(segments, 1):
                f.write(f"{i}. [{format_waktu(seg['start'])}] {seg['text'].strip()}\n")
        print(f"✅ Fallback text file created: {fallback_file}")

def transkrip_youtube_auto(video_url="https://youtu.be/l3X2SaCndiI?si=CfkAJHcFj0y254Pr", output_name="transkrip_hasil", base_folder="transcripts", model_size="large"):
    """
    Fungsi otomatis untuk membuat transkrip YouTube dengan format tabel yang indah
    Setiap run akan membuat folder output yang unik berdasarkan timestamp
    """
    try:
        print("🎬 YOUTUBE TRANSCRIPT GENERATOR (AUTO MODE)")
        print("Menggunakan YT-DLP + OpenAI Whisper AI")
        print("Setiap run akan membuat folder unik berdasarkan timestamp")
        print("=" * 65)
        
        # --- STEP 1: DOWNLOAD AUDIO ---
        print(f"\n📥 Step 1: Mengunduh audio dari YouTube...")
        print(f"🔗 URL: {video_url}")
        
        audio_file, video_info = download_audio_ytdlp(video_url, "temp_audio")
        
        if not audio_file or not video_info:
            error_msg = "Gagal mengunduh audio dari YouTube (mungkin video private, dihapus, atau region-blocked)"
            print(f"❌ {error_msg}")
            return {
                'status': 'error',
                'message': error_msg,
                'output_folder': None
            }
        
        print(f"✅ Audio berhasil diunduh!")
        print(f"📹 Judul: {video_info['title']}")
        print(f"⏱️ Durasi: {format_waktu(video_info['duration'])}")
        print(f"👤 Channel: {video_info['uploader']}")
        
        # --- CREATE UNIQUE OUTPUT FOLDER ---
        output_folder = create_unique_output_folder(video_info['title'], base_folder)        # --- STEP 2: LOAD AI MODEL ---
        print(f"\n🤖 Step 2: Memuat model Whisper AI...")
        
        # Pilihan model (dari kecil ke besar):
        # "tiny" (39MB) - Paling cepat, akurasi rendah
        # "base" (74MB) - Balance speed/accuracy 
        # "small" (244MB) - Akurasi lebih baik
        # "medium" (769MB) - Akurasi tinggi
        # "large" (1.5GB) - Akurasi tertinggi ⭐ RECOMMENDED
        
        print(f"📥 Loading Whisper model: {model_size}...")
        
        try:
            model = whisper.load_model(model_size)
            device = "cuda" if torch.cuda.is_available() else "cpu"
            print(f"✅ Model loaded successfully!")
            print(f"🖥️ Device: {device}")
            print(f"🧠 Model: Whisper {model_size.capitalize()}")
            if model_size == "large":
                print("🎯 Using highest accuracy model - perfect for training data!")
        except Exception as model_error:
            print(f"⚠️ Failed to load {model_size} model: {model_error}")
            print("🔄 Falling back to 'base' model...")
            model = whisper.load_model("base")
            device = "cuda" if torch.cuda.is_available() else "cpu"
            print(f"🖥️ Device: {device}")
            print(f"🧠 Model: Whisper Base (fallback)")
        
        # --- STEP 3: TRANSCRIBE ---
        print(f"\n🎙️ Step 3: Memproses transkrip...")
        print("⏳ Sedang menganalisis audio, harap tunggu...")
        
        result = model.transcribe(audio_file, fp16=(device=="cuda"))
        segments = result['segments']
        
        print(f"✅ Transkrip selesai!")
        print(f"📊 Total segmen: {len(segments)}")
        
        # Clean up
        if os.path.exists(audio_file):
            os.remove(audio_file)
            print("🗑️ File audio sementara telah dihapus")
          # --- STEP 4: CREATE BEAUTIFUL OUTPUT ---
        print(f"\n📋 Step 4: Membuat file output dengan format tabel profesional...")
        print(f"📁 Output folder: {output_folder}")
        
        # Create paragraphs first (needed for Excel)
        print("📝 Creating paragraph format...")
        paragraphs = create_paragraph_format(segments)
        
        # 1. Beautiful Excel Table (dengan sheet paragraf)
        excel_filename = f"{output_name}.xlsx"
        try:
            create_beautiful_table_excel(video_info, segments, paragraphs, excel_filename, output_folder)
            print(f"✅ Excel file created: {excel_filename}")
        except Exception as excel_error:
            print(f"⚠️ Excel creation failed: {excel_error}")
            print("📝 Continuing with other output formats...")
          # 2. Text format dengan numbering yang rapi  
        txt_filename = f"{output_name}.txt"
        txt_path = os.path.join(output_folder, txt_filename)
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write("=" * 100 + "\n")
            f.write(f"TRANSKRIP VIDEO YOUTUBE (SEGMEN)\n")
            f.write("=" * 100 + "\n")
            f.write(f"Judul    : {video_info['title']}\n")
            f.write(f"Channel  : {video_info['uploader']}\n")
            f.write(f"Durasi   : {format_waktu(video_info['duration'])}\n")
            f.write(f"Segmen   : {len(segments)} bagian\n")
            f.write(f"Tanggal  : {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Output   : {output_folder}\n")
            f.write("=" * 100 + "\n\n")
            
            for i, segment in enumerate(segments, 1):
                timestamp = format_waktu(segment['start'])
                text = segment['text'].strip()
                duration = segment['end'] - segment['start']
                word_count = len(text.split())
                
                f.write(f"[{i:3d}] [{timestamp}] ({duration:.1f}s) ({word_count} kata)\n")
                f.write(f"      {text}\n\n")
        
        # 2b. Format paragraf files (paragraphs already created above)
        paragraf_path, kontinyu_path = create_paragraph_output(video_info, paragraphs, txt_filename, output_folder)
        
        # 3. CSV untuk analisis data
        csv_filename = f"{output_name}.csv"
        csv_path = os.path.join(output_folder, csv_filename)
        df_data = []
        for i, segment in enumerate(segments, 1):
            df_data.append({
                'No': i,
                'Waktu_Mulai': format_waktu(segment['start']),
                'Waktu_Selesai': format_waktu(segment['end']),
                'Detik_Mulai': round(segment['start'], 2),
                'Detik_Selesai': round(segment['end'], 2),
                'Durasi_Detik': round(segment['end'] - segment['start'], 2),
                'Jumlah_Kata': len(segment['text'].strip().split()),
                'Jumlah_Karakter': len(segment['text'].strip()),
                'Teks': segment['text'].strip()
            })
        
        df = pd.DataFrame(df_data)
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        
        # 4. SRT untuk subtitle
        srt_filename = f"{output_name}.srt"
        srt_path = os.path.join(output_folder, srt_filename)
        with open(srt_path, 'w', encoding='utf-8') as f:
            for i, segment in enumerate(segments, 1):
                start_time = format_srt_time(segment['start'])
                end_time = format_srt_time(segment['end'])
                text = segment['text'].strip()
                
                f.write(f"{i}\n")
                f.write(f"{start_time} --> {end_time}\n")
                f.write(f"{text}\n\n")
        
        # 5. Create info file with folder and run details
        info_filename = "run_info.txt"
        info_path = os.path.join(output_folder, info_filename)
        with open(info_path, 'w', encoding='utf-8') as f:
            f.write("INFORMASI TRANSKRIP\n")
            f.write("=" * 50 + "\n")
            f.write(f"Tanggal Run    : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Video URL      : {video_url}\n")
            f.write(f"Video Judul    : {video_info['title']}\n")
            f.write(f"Channel        : {video_info['uploader']}\n")
            f.write(f"Durasi         : {format_waktu(video_info['duration'])}\n")
            f.write(f"Output Folder  : {output_folder}\n")
            f.write(f"Total Segmen   : {len(segments)}\n")
            f.write("=" * 50 + "\n")
            f.write("File yang dihasilkan:\n")
            f.write(f"- {excel_filename} (3 sheets: Segmen, Paragraf, Statistik)\n")
            f.write(f"- {txt_filename} (format segmen)\n")
            f.write(f"- {os.path.basename(paragraf_path)} (format paragraf)\n")
            f.write(f"- {os.path.basename(kontinyu_path)} (format kontinyu)\n")
            f.write(f"- {csv_filename}\n")
            f.write(f"- {srt_filename}\n")
            f.write(f"- {info_filename}\n")
          # Calculate final statistics
        total_words = sum(len(seg['text'].split()) for seg in segments)
        total_chars = sum(len(seg['text']) for seg in segments)
        words_per_minute = (total_words / (video_info['duration'] / 60)) if video_info['duration'] > 0 else 0
        
        print("\n🎉 SELESAI! File berhasil dibuat di folder:")
        print(f"📁 {output_folder}")
        print("\n📄 File yang dibuat:")
        print(f"📊 Excel (3 Sheets)          : {excel_filename}")
        print(f"   🔸 Sheet 1: Segmen detail")
        print(f"   🔸 Sheet 2: Format paragraf")  
        print(f"   🔸 Sheet 3: Statistik lengkap")
        print(f"📝 Text (Format Segmen)      : {txt_filename}")
        print(f"📄 Text (Format Paragraf)    : {os.path.basename(paragraf_path)}")
        print(f"📖 Text (Format Kontinyu)    : {os.path.basename(kontinyu_path)}")
        print(f"📋 CSV (Data Analysis)       : {csv_filename}")
        print(f"🎬 SRT (Subtitle)            : {srt_filename}")
        print(f"ℹ️ Info Run                  : {info_filename}")
        print("\n📈 STATISTIK:")
        print(f"   • Total segmen    : {len(segments)}")
        print(f"   • Total paragraf  : {len(paragraphs)}")
        print(f"   • Total kata      : {total_words:,}")
        print(f"   • Total karakter  : {total_chars:,}")
        print(f"   • Kecepatan bicara: {words_per_minute:.1f} kata/menit")
        print("=" * 65)
        
        return {
            'status': 'success',
            'output_folder': output_folder,            'files': {
                'excel': os.path.join(output_folder, excel_filename),
                'txt': txt_path,
                'txt_paragraf': paragraf_path,
                'txt_kontinyu': kontinyu_path,
                'csv': csv_path,
                'srt': srt_path,
                'info': info_path
            },            'stats': {
                'segments': len(segments),
                'paragraphs': len(paragraphs),
                'words': total_words,
                'characters': total_chars,
                'duration': video_info['duration'],
                'words_per_minute': words_per_minute
            },
            'video_info': video_info
        }
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return {'status': 'error', 'message': str(e)}

def format_srt_time(seconds):
    """Format waktu untuk file SRT"""
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    millisecs = int((seconds - int(seconds)) * 1000)
    return f"{hours:02d}:{minutes:02d}:{secs:02d},{millisecs:03d}"

# --- PARAGRAPH CREATION FUNCTIONS ---
def create_paragraph_format(segments, min_paragraph_duration=30, max_paragraph_duration=120):
    """
    Menggabungkan segmen menjadi paragraf yang lebih natural
    """
    if not segments:
        return []
    
    paragraphs = []
    current_paragraph = {
        'start': segments[0]['start'],
        'text': '',
        'segments': []
    }
    
    for segment in segments:
        # Tambahkan segmen ke paragraf saat ini
        current_paragraph['text'] += segment['text'].strip() + ' '
        current_paragraph['segments'].append(segment)
        
        # Hitung durasi paragraf saat ini
        duration = segment['end'] - current_paragraph['start']
        
        # Kondisi untuk memulai paragraf baru:
        # 1. Durasi sudah mencapai minimum DAN ada jeda panjang
        # 2. Durasi sudah mencapai maksimum
        # 3. Deteksi akhir kalimat dengan tanda baca
        text = segment['text'].strip()
        has_sentence_end = text.endswith(('.', '!', '?', '。', '！', '？'))
        
        should_break = (
            (duration >= min_paragraph_duration and has_sentence_end) or
            duration >= max_paragraph_duration
        )
        
        if should_break:
            # Selesaikan paragraf saat ini
            current_paragraph['end'] = segment['end']
            current_paragraph['text'] = current_paragraph['text'].strip()
            current_paragraph['duration'] = current_paragraph['end'] - current_paragraph['start']
            current_paragraph['word_count'] = len(current_paragraph['text'].split())
            
            paragraphs.append(current_paragraph)
            
            # Mulai paragraf baru
            current_paragraph = {
                'start': segment['end'],
                'text': '',
                'segments': []
            }
    # Tambahkan paragraf terakhir jika ada
    if current_paragraph['text'].strip():
        current_paragraph['end'] = segments[-1]['end']
        current_paragraph['text'] = current_paragraph['text'].strip()
        current_paragraph['duration'] = current_paragraph['end'] - current_paragraph['start']
        current_paragraph['word_count'] = len(current_paragraph['text'].split())
        paragraphs.append(current_paragraph)
    
    return paragraphs

def extract_playlist_urls(playlist_url):
    """
    Extract individual video URLs from YouTube playlist
    """
    try:
        ydl_opts = {
            'quiet': True,
            'extract_flat': True,  # Don't download, just get URLs
        }
        
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            playlist_info = ydl.extract_info(playlist_url, download=False)
            
            if 'entries' not in playlist_info:
                return None, "Not a valid playlist"
            
            video_urls = []
            playlist_title = playlist_info.get('title', 'Unknown Playlist')
            
            for entry in playlist_info['entries']:
                if entry:  # Some entries might be None
                    video_url = f"https://www.youtube.com/watch?v={entry['id']}"
                    video_urls.append({
                        'url': video_url,
                        'title': entry.get('title', 'Unknown'),
                        'id': entry['id']
                    })
            
            return {
                'playlist_title': playlist_title,
                'total_videos': len(video_urls),
                'videos': video_urls
            }, None
            
    except Exception as e:
        return None, str(e)

def is_playlist_url(url):
    """Check if URL is a playlist"""
    return 'playlist' in url or 'list=' in url

def create_paragraph_output(video_info, paragraphs, filename, output_folder="."):
    """
    Membuat file output dalam format paragraf yang rapi
    """
    # Paragraf TXT format
    paragraf_path = os.path.join(output_folder, filename.replace('.txt', '_paragraf.txt'))
    
    with open(paragraf_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("TRANSKRIP VIDEO YOUTUBE (FORMAT PARAGRAF)\n")
        f.write("=" * 80 + "\n")
        f.write(f"Judul    : {video_info['title']}\n")
        f.write(f"Channel  : {video_info['uploader']}\n")
        f.write(f"Durasi   : {format_waktu(video_info['duration'])}\n")
        f.write(f"Paragraf : {len(paragraphs)} bagian\n")
        f.write(f"Tanggal  : {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        
        for i, paragraph in enumerate(paragraphs, 1):
            start_time = format_waktu(paragraph['start'])
            end_time = format_waktu(paragraph['end'])
            duration = paragraph['duration']
            word_count = paragraph['word_count']
            
            f.write(f"PARAGRAF {i} [{start_time} - {end_time}] ({duration:.1f}s, {word_count} kata)\n")
            f.write("-" * 60 + "\n")
            f.write(f"{paragraph['text']}\n\n")
    
    # Paragraf kontinyu (tanpa timestamp)
    kontinyu_path = os.path.join(output_folder, filename.replace('.txt', '_kontinyu.txt'))
    
    with open(kontinyu_path, 'w', encoding='utf-8') as f:
        f.write(f"TRANSKRIP: {video_info['title']}\n")
        f.write(f"Channel: {video_info['uploader']}\n")
        f.write(f"Durasi: {format_waktu(video_info['duration'])}\n")
        f.write("=" * 60 + "\n\n")
        
        for paragraph in paragraphs:
            f.write(f"{paragraph['text']}\n\n")
    
    return paragraf_path, kontinyu_path

def transkrip_playlist_auto(playlist_url, output_name="playlist_transkrip", base_folder="transcripts", model_size="large"):
    """
    Proses playlist YouTube secara batch
    """
    print("🎵 YOUTUBE PLAYLIST TRANSCRIPT GENERATOR")
    print("=" * 60)
    
    # Extract playlist info
    playlist_info, error = extract_playlist_urls(playlist_url)
    if error:
        print(f"❌ Error extracting playlist: {error}")
        return {'status': 'error', 'message': error}
    
    print(f"📋 Playlist: {playlist_info['playlist_title']}")
    print(f"🎬 Total videos: {playlist_info['total_videos']}")
    print()
    
    # Create playlist output folder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    playlist_folder = os.path.join(base_folder, f"PLAYLIST_{timestamp}_{playlist_info['playlist_title'][:30]}")
    os.makedirs(playlist_folder, exist_ok=True)
    print(f"📁 Playlist folder: {playlist_folder}")
    
    results = {
        'status': 'success',
        'playlist_info': playlist_info,
        'playlist_folder': playlist_folder,
        'processed_videos': [],
        'failed_videos': [],
        'total_processed': 0
    }
    
    # Process each video
    for i, video in enumerate(playlist_info['videos'], 1):
        print(f"\n{'='*60}")
        print(f"🎬 Processing video {i}/{playlist_info['total_videos']}")
        print(f"📹 {video['title']}")
        print(f"🔗 {video['url']}")
        
        try:
            # Create subfolder for this video
            safe_title = re.sub(r'[^a-zA-Z0-9\s]', '', video['title'])[:30]
            video_folder = os.path.join(playlist_folder, f"{i:03d}_{safe_title}")
              # Process single video
            result = transkrip_youtube_auto(
                video_url=video['url'],
                output_name=f"video_{i:03d}_{output_name}",
                base_folder=video_folder,
                model_size=model_size
            )
            
            if result['status'] == 'success':
                results['processed_videos'].append({
                    'index': i,
                    'title': video['title'],
                    'url': video['url'],
                    'output_folder': result['output_folder'],
                    'stats': result['stats']
                })
                print(f"✅ Video {i} completed successfully!")
            else:
                results['failed_videos'].append({
                    'index': i,
                    'title': video['title'],
                    'url': video['url'],
                    'error': result.get('message', 'Unknown error')
                })
                print(f"⚠️ Video {i} skipped: {result.get('message', 'Unknown error')}")
                print(f"⏭️ Continuing to next video...")
                
        except Exception as e:
            results['failed_videos'].append({
                'index': i,
                'title': video['title'],
                'url': video['url'],
                'error': str(e)
            })
            print(f"⚠️ Video {i} failed with exception: {e}")
            print(f"⏭️ Continuing to next video...")
        
        results['total_processed'] = i
    
    # Create playlist summary
    summary_path = os.path.join(playlist_folder, "PLAYLIST_SUMMARY.txt")
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("YOUTUBE PLAYLIST TRANSCRIPT SUMMARY\n")
        f.write("=" * 50 + "\n")
        f.write(f"Playlist: {playlist_info['playlist_title']}\n")
        f.write(f"Processed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total Videos: {playlist_info['total_videos']}\n")
        f.write(f"Successful: {len(results['processed_videos'])}\n")
        f.write(f"Failed: {len(results['failed_videos'])}\n")
        f.write("=" * 50 + "\n\n")
        
        f.write("SUCCESSFUL VIDEOS:\n")
        f.write("-" * 30 + "\n")
        for video in results['processed_videos']:
            f.write(f"{video['index']:3d}. {video['title']}\n")
            f.write(f"     Folder: {video['output_folder']}\n")
            f.write(f"     Stats: {video['stats']['segments']} segments, {video['stats']['words']} words\n\n")
        
        if results['failed_videos']:
            f.write("FAILED VIDEOS:\n")
            f.write("-" * 30 + "\n")
            for video in results['failed_videos']:
                f.write(f"{video['index']:3d}. {video['title']}\n")
                f.write(f"     Error: {video['error']}\n\n")
    
    print(f"\n🎉 PLAYLIST PROCESSING COMPLETED!")
    print(f"✅ Successful: {len(results['processed_videos'])}/{playlist_info['total_videos']}")
    print(f"❌ Failed: {len(results['failed_videos'])}/{playlist_info['total_videos']}")
    print(f"📁 Output folder: {playlist_folder}")
    print(f"📄 Summary: {summary_path}")
    
    # === CREATE CONSOLIDATED EXCEL FOR TRAINING ===
    if results['processed_videos']:
        print(f"\n📊 Creating consolidated Excel for training...")
        consolidated_filename = f"PLAYLIST_CONSOLIDATED_{playlist_info['playlist_title'][:30]}.xlsx"
        consolidated_result = create_playlist_consolidated_excel(
            results, playlist_info, consolidated_filename, playlist_folder
        )
        
        if consolidated_result['status'] == 'success':
            print(f"✅ Consolidated Excel created: {consolidated_filename}")
            print(f"📊 Total words: {consolidated_result['stats']['total_words']:,}")
            print(f"📄 Total characters: {consolidated_result['stats']['total_characters']:,}")
            print(f"🤖 Estimated tokens: {consolidated_result['stats']['estimated_tokens']:,}")
            print("📋 Sheets: Ringkasan | Semua Transkrip | Teks Continuous | Statistik")
            results['consolidated_excel'] = consolidated_result
        else:
            print(f"⚠️ Failed to create consolidated Excel: {consolidated_result['message']}")
    
    return results

def create_playlist_consolidated_excel(playlist_results, playlist_info, filename, output_folder="."):
    """
    Membuat satu file Excel gabungan untuk seluruh playlist
    Fokus pada teks untuk pelatihan, tanpa detail timestamp
    """
    try:
        # Ensure output folder exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder, exist_ok=True)
        
        full_filename = os.path.join(output_folder, filename)
        print(f"📊 Creating consolidated playlist Excel: {full_filename}")
        
        wb = Workbook()
        
        # === SHEET 1: RINGKASAN PLAYLIST ===
        ws_summary = wb.active
        ws_summary.title = "Ringkasan Playlist"
        
        # Styling
        primary_blue = "1F4E79"
        light_blue = "D9E2F3"
        accent_blue = "2E75B6"
        white = "FFFFFF"
        light_gray = "F8F9FA"
        green = "70AD47"
        
        title_font = Font(name='Calibri', size=16, bold=True, color=primary_blue)
        header_font = Font(name='Calibri', size=11, bold=True, color=white)
        data_font = Font(name='Calibri', size=10)
        
        header_fill = PatternFill(start_color=primary_blue, end_color=primary_blue, fill_type="solid")
        light_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
        
        thick_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Title
        title_cell = ws_summary.cell(row=1, column=1, value="TRANSKRIP PLAYLIST YOUTUBE - RINGKASAN")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells('A1:F1')
        
        # Playlist info
        playlist_title_cell = ws_summary.cell(row=2, column=1, value=playlist_info['playlist_title'])
        playlist_title_cell.font = Font(name='Calibri', size=12, bold=True)
        playlist_title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_summary.merge_cells('A2:F2')
        
        # Headers
        header_row = 4
        headers = ["No", "Judul Video", "Durasi", "Total Kata", "Total Paragraf", "Status"]
        column_widths = [6, 50, 12, 12, 15, 12]
        
        for col, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws_summary.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            ws_summary.column_dimensions[chr(64 + col)].width = width
        
        # Data rows
        data_start_row = header_row + 1
        total_words = 0
        total_paragraphs = 0
        
        for i, video in enumerate(playlist_results['processed_videos']):
            row = data_start_row + i
            fill = light_fill if i % 2 == 0 else PatternFill(start_color=white, end_color=white, fill_type="solid")
            
            # Data
            ws_summary.cell(row=row, column=1, value=video['index']).font = data_font
            ws_summary.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row, column=1).border = thin_border
            ws_summary.cell(row=row, column=1).fill = fill
            
            ws_summary.cell(row=row, column=2, value=video['title']).font = data_font
            ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal="left", wrap_text=True)
            ws_summary.cell(row=row, column=2).border = thin_border
            ws_summary.cell(row=row, column=2).fill = fill
            
            ws_summary.cell(row=row, column=3, value=format_waktu(video['stats']['duration'])).font = data_font
            ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row, column=3).border = thin_border
            ws_summary.cell(row=row, column=3).fill = fill
            
            ws_summary.cell(row=row, column=4, value=f"{video['stats']['words']:,}").font = data_font
            ws_summary.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row, column=4).border = thin_border
            ws_summary.cell(row=row, column=4).fill = fill
            
            ws_summary.cell(row=row, column=5, value=f"{video['stats']['paragraphs']:,}").font = data_font
            ws_summary.cell(row=row, column=5).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row, column=5).border = thin_border
            ws_summary.cell(row=row, column=5).fill = fill
            
            ws_summary.cell(row=row, column=6, value="✅ Success").font = Font(name='Calibri', size=10, color=green)
            ws_summary.cell(row=row, column=6).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=row, column=6).border = thin_border
            ws_summary.cell(row=row, column=6).fill = fill
            
            total_words += video['stats']['words']
            total_paragraphs += video['stats']['paragraphs']
        
        # Summary row
        summary_row = data_start_row + len(playlist_results['processed_videos']) + 1
        ws_summary.cell(row=summary_row, column=1, value="TOTAL").font = Font(name='Calibri', size=11, bold=True)
        ws_summary.cell(row=summary_row, column=4, value=f"{total_words:,}").font = Font(name='Calibri', size=11, bold=True)
        ws_summary.cell(row=summary_row, column=5, value=f"{total_paragraphs:,}").font = Font(name='Calibri', size=11, bold=True)
        
        # === SHEET 2: SEMUA TRANSKRIP GABUNGAN ===
        ws_all = wb.create_sheet("Semua Transkrip")
        
        # Title
        title_cell = ws_all.cell(row=1, column=1, value="SEMUA TRANSKRIP PLAYLIST (UNTUK PELATIHAN TEKS)")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_all.merge_cells('A1:C1')
        
        # Playlist info
        playlist_title_cell = ws_all.cell(row=2, column=1, value=playlist_info['playlist_title'])
        playlist_title_cell.font = Font(name='Calibri', size=12, bold=True)
        playlist_title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_all.merge_cells('A2:C2')
        
        # Headers
        header_row = 4
        headers = ["No Video", "Judul Video", "Teks Lengkap (Gabungan Semua Paragraf)"]
        column_widths = [10, 40, 100]
        
        for col, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws_all.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
            ws_all.column_dimensions[chr(64 + col)].width = width
        
        # Ambil data transkrip dari file yang sudah dibuat
        current_row = header_row + 1
        all_texts = []
        
        for video in playlist_results['processed_videos']:
            try:
                # Cari file paragraf dari folder output video
                video_folder = video['output_folder']
                paragraf_files = []
                
                # Cari file yang berakhiran '_kontinyu.txt' (teks bersih tanpa timestamp)
                for file in os.listdir(video_folder):
                    if file.endswith('_kontinyu.txt'):
                        paragraf_files.append(os.path.join(video_folder, file))
                
                if paragraf_files:
                    # Baca file kontinyu pertama yang ditemukan
                    with open(paragraf_files[0], 'r', encoding='utf-8') as f:
                        content = f.read()
                        
                        # Extract teks bersih (skip header)
                        lines = content.split('\n')
                        text_start = False
                        clean_text = []
                        
                        for line in lines:
                            if '=' in line and not text_start:
                                text_start = True
                                continue
                            elif text_start and line.strip():
                                if not any(x in line for x in ['TRANSKRIP:', 'Channel:', 'Durasi:', '=']):
                                    clean_text.append(line.strip())
                        
                        # Gabungkan menjadi satu teks
                        full_text = ' '.join(clean_text).strip()
                        all_texts.append(full_text)
                        
                        if full_text:
                            fill = light_fill if (current_row - header_row - 1) % 2 == 0 else PatternFill(start_color=white, end_color=white, fill_type="solid")
                            
                            # No Video
                            cell = ws_all.cell(row=current_row, column=1, value=video['index'])
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border
                            cell.fill = fill
                            
                            # Judul Video
                            cell = ws_all.cell(row=current_row, column=2, value=video['title'])
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            cell.border = thin_border
                            cell.fill = fill
                            
                            # Teks Lengkap
                            cell = ws_all.cell(row=current_row, column=3, value=full_text)
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            cell.border = thin_border
                            cell.fill = fill
                            
                            # Set row height based on text length
                            estimated_lines = max(3, len(full_text) // 150)
                            ws_all.row_dimensions[current_row].height = max(60, estimated_lines * 20)
                            

                            current_row += 1
                
            except Exception as e:
                print(f"⚠️ Warning: Could not read transcript for video {video['index']}: {e}")
                continue
        
        # === SHEET 3: TEKS GABUNGAN CONTINUOUS ===
        ws_continuous = wb.create_sheet("Teks Continuous")
        
        # Title
        title_cell = ws_continuous.cell(row=1, column=1, value="TEKS GABUNGAN SELURUH PLAYLIST")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_continuous.merge_cells('A1:A1')
        
        # Info
        info_cell = ws_continuous.cell(row=2, column=1, value="(Semua transkrip digabung menjadi satu teks panjang untuk pelatihan)")
        info_cell.font = Font(name='Calibri', size=10, italic=True)
        info_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Gabungkan semua teks
        combined_text = '\n\n'.join(all_texts)
        
        # Tulis teks gabungan
        cell = ws_continuous.cell(row=4, column=1, value=combined_text)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_continuous.column_dimensions['A'].width = 120
        
        # Set tinggi row untuk teks panjang
        estimated_lines = len(combined_text) // 120
        ws_continuous.row_dimensions[4].height = max(500, estimated_lines * 15)
        
        # === SHEET 4: STATISTIK PLAYLIST ===
        ws_stats = wb.create_sheet("Statistik Playlist")
        
        # Calculate comprehensive stats
        total_duration = sum(v['stats']['duration'] for v in playlist_results['processed_videos'])
        total_segments = sum(v['stats']['segments'] for v in playlist_results['processed_videos'])
        avg_words_per_minute = (total_words / (total_duration / 60)) if total_duration > 0 else 0
        
        stats_data = [
            ["STATISTIK LENGKAP PLAYLIST", ""],
            ["", ""],
            ["Informasi Playlist", ""],
            ["Nama Playlist", playlist_info['playlist_title']],
            ["Total Video", f"{len(playlist_results['processed_videos'])}"],
            ["Video Berhasil", f"{len(playlist_results['processed_videos'])}"],
            ["Video Gagal", f"{len(playlist_results['failed_videos'])}"],
            ["", ""],
            ["Statistik Konten", ""],
            ["Total Durasi", format_waktu(total_duration)],
            ["Total Kata", f"{total_words:,}"],
            ["Total Paragraf", f"{total_paragraphs:,}"],
            ["Total Segmen", f"{total_segments:,}"],
            ["Rata-rata Kata/Menit", f"{avg_words_per_minute:.1f}"],
            ["Rata-rata Kata/Video", f"{total_words/len(playlist_results['processed_videos']):.0f}"],
            ["Rata-rata Paragraf/Video", f"{total_paragraphs/len(playlist_results['processed_videos']):.1f}"],
            ["", ""],
            ["Info Pelatihan", ""],
            ["Panjang Teks Gabungan", f"{len(combined_text):,} karakter"],
            ["Estimasi Token (GPT)", f"{len(combined_text)//4:,}"],
            ["Cocok untuk", "Training, Fine-tuning, Dataset"],
        ]
        
        # Write statistics
        for row, (label, value) in enumerate(stats_data, 1):
            if row == 1:  # Title
                cell = ws_stats.cell(row=row, column=1, value=label)
                cell.font = Font(name='Calibri', size=16, bold=True, color=primary_blue)
                cell.alignment = Alignment(horizontal="center")
                ws_stats.merge_cells(f'A{row}:B{row}')
            elif label and not any(x in label for x in ["Informasi", "Statistik", "Info"]):  # Data
                # Label
                label_cell = ws_stats.cell(row=row, column=1, value=label)
                label_cell.font = Font(name='Calibri', size=11, bold=True)
                label_cell.alignment = Alignment(horizontal="right")
                
                # Value
                value_cell = ws_stats.cell(row=row, column=2, value=value)
                value_cell.font = Font(name='Calibri', size=11)
                value_cell.alignment = Alignment(horizontal="left")
            elif label:  # Section headers
                cell = ws_stats.cell(row=row, column=1, value=label)
                cell.font = Font(name='Calibri', size=12, bold=True, color=accent_blue)
                ws_stats.merge_cells(f'A{row}:B{row}')
        
        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 30
        
        # Save workbook
        wb.save(full_filename)
        print(f"✅ Consolidated playlist Excel saved: {full_filename}")
        
        return {
            'status': 'success',
            'file_path': full_filename,
            'stats': {
                'total_videos': len(playlist_results['processed_videos']),
                'total_words': total_words,
                'total_paragraphs': total_paragraphs,
                'total_characters': len(combined_text),
                'estimated_tokens': len(combined_text)//4
            }
        }
        
    except Exception as e:
        print(f"❌ Error creating consolidated Excel: {e}")
        return {'status': 'error', 'message': str(e)}

# --- MODEL SELECTION ---
def choose_whisper_model():
    """
    Fungsi untuk memilih model Whisper berdasarkan kebutuhan user
    """
    print("\n🤖 PILIHAN MODEL WHISPER AI")
    print("=" * 50)
    print("1. tiny   (39MB)  - ⚡⚡⚡⚡⚡ Paling cepat, akurasi ⭐⭐")
    print("2. base   (74MB)  - ⚡⚡⚡⚡ Cepat, akurasi ⭐⭐⭐")
    print("3. small  (244MB) - ⚡⚡⚡ Sedang, akurasi ⭐⭐⭐⭐")
    print("4. medium (769MB) - ⚡⚡ Lambat, akurasi ⭐⭐⭐⭐⭐")
    print("5. large  (1.5GB) - ⚡ Paling lambat, akurasi ⭐⭐⭐⭐⭐⭐ 🏆")
    print("\n💡 Rekomendasi:")
    print("   • Testing/preview: tiny atau base")
    print("   • Produksi normal: small atau medium") 
    print("   • Training AI/dataset: large (terbaik)")
    print("   • GPU lemah/RAM sedikit: base")
    
    choice = input("\nPilih model (1-5) atau Enter untuk default (large): ").strip()
    
    model_map = {
        '1': 'tiny',
        '2': 'base', 
        '3': 'small',
        '4': 'medium',
        '5': 'large',
        '': 'large'  # Default
    }
    
    selected_model = model_map.get(choice, 'large')
    
    print(f"✅ Model dipilih: {selected_model}")
    if selected_model == 'large':
        print("🎯 Excellent choice! Model terbaik untuk akurasi maksimal.")
    elif selected_model in ['medium', 'small']:
        print("👍 Good choice! Balance yang baik antara akurasi dan kecepatan.")
    else:
        print("⚡ Fast choice! Cocok untuk testing atau hardware terbatas.")
    
    return selected_model

# --- MAIN EXECUTION ---
if __name__ == "__main__":
    print("🚀 YOUTUBE TRANSCRIPT GENERATOR WITH PARAGRAPH & PLAYLIST SUPPORT")
    print("📁 Features: Segmen + Paragraf + Kontinyu + Playlist Support")
    print("=" * 70)
    
    # Model selection
    selected_model = choose_whisper_model()
    
    # URL input
    url_input = input("\nMasukkan URL YouTube (video/playlist) atau Enter untuk default: ").strip()
    if url_input:
        video_url = url_input
    else:
        video_url = "https://youtu.be/l3X2SaCndiI?si=CfkAJHcFj0y254Pr"
      # Check if it's a playlist
    is_playlist = is_playlist_url(video_url)
    
    if is_playlist:
        print("🎵 PLAYLIST DETECTED!")
        print("📋 This will process all videos in the playlist")
        print("🆕 BONUS: Excel konsolidasi otomatis untuk pelatihan teks!")
        print("   📊 4 sheets: Ringkasan | Semua Transkrip | Teks Continuous | Statistik")
        print("   🎯 Perfect untuk training AI & review cepat")
        confirm = input("Continue with playlist processing? (y/n): ").strip().lower()
        if confirm not in ['y', 'yes', '']:
            print("❌ Playlist processing cancelled.")
            exit()
        
        # Nama output file untuk playlist
        output_input = input("Masukkan nama base untuk file output (atau Enter untuk default): ").strip()
        if output_input:
            output_name = output_input
        else:
            output_name = "transkrip_playlist"
        
        # Folder base (opsional)
        folder_input = input("Masukkan nama folder base (atau Enter untuk 'transcripts'): ").strip()
        if folder_input:
            base_folder = folder_input
        else:
            base_folder = "transcripts"
        
        print(f"\n🎯 Processing playlist: {video_url}")
        print(f"📁 Base folder: {base_folder}")
        print(f"📄 Output name: {output_name}")
        print("\n🔄 Starting playlist processing...")
          # Process playlist
        result = transkrip_playlist_auto(video_url, output_name, base_folder, selected_model)
        
        if result['status'] == 'success':
            print(f"\n✨ PLAYLIST COMPLETED!")
            print(f"📋 Playlist: {result['playlist_info']['playlist_title']}")
            print(f"✅ Processed: {len(result['processed_videos'])}/{result['playlist_info']['total_videos']} videos")
            print(f"📁 Output folder: {result['playlist_folder']}")
            
            # Create consolidated Excel
            excel_consolidated_filename = f"CONSOLIDATED_{output_name}.xlsx"
            create_playlist_consolidated_excel(result, result['playlist_info'], excel_consolidated_filename, result['playlist_folder'])
        else:
            print(f"\n❌ PLAYLIST FAILED: {result['message']}")
    
    else:
        print("🎬 SINGLE VIDEO DETECTED")
        
        # Nama output file
        output_input = input("Masukkan nama file output (atau Enter untuk default): ").strip()
        if output_input:
            output_name = output_input
        else:
            output_name = "transkrip_youtube_professional"
        
        # Folder base (opsional)
        folder_input = input("Masukkan nama folder base (atau Enter untuk 'transcripts'): ").strip()
        if folder_input:
            base_folder = folder_input
        else:
            base_folder = "transcripts"
        
        print(f"\n🎯 Processing video: {video_url}")
        print(f"📁 Base folder: {base_folder}")
        print(f"📄 Output name: {output_name}")
        print("💡 Output akan include: Segmen + Paragraf + Kontinyu + Excel + CSV + SRT")
          # Process single video
        result = transkrip_youtube_auto(video_url, output_name, base_folder, selected_model)
        
        if result['status'] == 'success':
            print(f"\n✨ BERHASIL!")
            print(f"Video '{result['video_info']['title']}' telah berhasil ditranskrip.")
            print(f"📁 Silakan cek folder: {result['output_folder']}")
            print("🎉 Setiap run membuat folder baru dengan timestamp unik!")
        else:
            print(f"\n❌ GAGAL: {result['message']}")
            print("💡 Pastikan URL valid dan koneksi internet stabil.")
